# -*- coding: utf-8 -*-
"""Gotion 表格管理 - CSV / Excel 导入处理函数"""

from typing import Any, List

from fastapi import HTTPException, UploadFile, File, Form

from src.db_manage.database import DatabaseManager
from src.db_manage.models import GotionTableModel, GotionColumnModel, GotionRowModel
from .tables_handler import _slug


async def import_from_file(table_id: int, file: UploadFile = File(...), has_header: bool = Form(True)):
    """从 CSV 或 Excel 导入数据到指定表"""
    import csv
    import io

    table = GotionTableModel.find_by_id(id=table_id)
    if not table:
        raise HTTPException(404, '表不存在')

    if not file:
        raise HTTPException(400, '没有提供文件')

    # 读取文件内容
    file_content = await file.read()
    filename = file.filename or ''

    # 判断文件类型
    is_excel = filename.lower().endswith(('.xlsx', '.xls'))

    headers: List[str] = []
    rows_data: List[List[Any]] = []

    try:
        if is_excel:
            # Excel 文件解析
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise HTTPException(500, '服务端缺少 openpyxl 依赖，无法解析 Excel，请先 pip install openpyxl')
            wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
            ws = wb.active

            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                raise HTTPException(400, '文件为空')

            if has_header:
                headers = [str(c) if c is not None else f'列{i+1}' for i, c in enumerate(all_rows[0])]
                rows_data = [list(row) for row in all_rows[1:]]
            else:
                # 自动生成列名
                first_row = all_rows[0]
                headers = [f'列{i+1}' for i in range(len(first_row))]
                rows_data = [list(row) for row in all_rows]
            wb.close()
        else:
            # CSV 文件解析（优先 UTF-8/BOM，兼容 Excel 导出的 GBK 编码）
            try:
                text = file_content.decode('utf-8-sig')
            except UnicodeDecodeError:
                text = file_content.decode('gb18030')
            reader = csv.reader(io.StringIO(text))
            all_rows = list(reader)

            if not all_rows:
                raise HTTPException(400, '文件为空')

            if has_header:
                headers = [str(c).strip() or f'列{i+1}' for i, c in enumerate(all_rows[0])]
                rows_data = all_rows[1:]
            else:
                headers = [f'列{i+1}' for i in range(len(all_rows[0]))]
                rows_data = all_rows
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f'文件解析失败: {str(e)}')

    # 列匹配：表头 slug 与现有列 key 相同 → 复用该列；否则建新列。
    # 同一文件里的重复表头依次加后缀，避免互相覆盖。
    existing_keys = {c.key for c in GotionColumnModel.find_by_table(table_id)}
    col_map = {}        # 列索引 -> key
    used_keys = set()   # 本次导入已占用的 key
    new_columns = []    # 待创建的 (表头名, key)

    for i, header in enumerate(headers):
        key = _slug(header)
        if key in existing_keys and key not in used_keys:
            col_map[i] = key
            used_keys.add(key)
            continue
        base_key, counter = key, 1
        while key in existing_keys or key in used_keys:
            key = f'{base_key}_{counter}'
            counter += 1
        col_map[i] = key
        used_keys.add(key)
        new_columns.append((header, key))

    # 新列 + 全部行在一个事务中写入，失败整体回滚，不留半截导入
    new_cols_created = 0
    rows_imported = 0
    with DatabaseManager().transaction():
        next_col_sort = GotionColumnModel.get_next_sort_order(table_id)
        for header, key in new_columns:
            new_col = GotionColumnModel(
                table_id=table_id,
                name=header,
                key=key,
                type='text',
                config='{}',
                width=200,
                sort_order=next_col_sort,
            )
            if not new_col.save():
                raise HTTPException(500, f'创建列「{header}」失败')
            next_col_sort += 1
            new_cols_created += 1

        next_row_sort = GotionRowModel.get_next_sort_order(table_id)
        for row_values in rows_data:
            if not any(cell is not None and str(cell).strip() != '' for cell in row_values):
                continue  # 跳过空行

            data = {}
            for i, value in enumerate(row_values):
                if i in col_map:
                    data[col_map[i]] = '' if value is None else str(value).strip()

            row = GotionRowModel(table_id=table_id, data=data, sort_order=next_row_sort)
            if not row.save():
                raise HTTPException(500, f'导入第 {rows_imported + 1} 行失败')
            next_row_sort += 1
            rows_imported += 1

    return {
        'message': f'成功导入 {rows_imported} 行数据',
        'rows_imported': rows_imported,
        'new_columns_created': new_cols_created,
    }
